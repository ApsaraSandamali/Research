import openpyxl
import os

# Create a new Excel workbook and select the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Add column headers to the worksheet
column_names = ['problem instance No', 'Cost for OR', 'Time for OR', 'Time for HTN', 'Step size for HTN']
for col, name in enumerate(column_names, start=1):
    worksheet.cell(row=1, column=col, value=name)

# Loop through all input files and write the data to the worksheet
row = 2
for i in range(19):
    with open(f'{os.getcwd()}/or_output/output{i}.txt', 'r') as file:
        total_cost = None
        total_time = None

        # Search for the total cost and total time in each line of the file
        for line in file:
            if 'Total cost:' in line:
                total_cost = line.strip().split(':')[-1].strip()
            elif 'Total runtime:' in line:
                total_time = line.strip().split(':')[-1].strip()

            # If both values are found, break out of the loop
            if total_cost is not None and total_time is not None:
                break

    with open(f'{os.getcwd()}/htn_output/output{i}.txt', 'r') as file:
        htn_total_time = None

        # Search for the total cost and total time in each line of the file
        for line in file:
            if 'Total time:' in line:
                htn_total_time = line.strip().split(':')[-1].strip()

            # If both values are found, break out of the loop
            if htn_total_time is not None:
                break

    with open(f'{os.getcwd()}/htn_output/output{i}.txt', 'r') as file:
        maxNum = None
        minNum = None
        lines = file.readlines()

        for j, line in enumerate(lines):
            if 'Planning' in line:
                line = lines[j-1]
                minNum = line.split()[0]

            if minNum is not None:
                break

        for j, line in enumerate(lines):
            if 'root' in line:
                line = lines[j-1]
                maxNum = line.split()[0]

            if maxNum is not None:
                break

        step_size = int(maxNum) - int(minNum)

    # Write the data to the worksheet
    worksheet.cell(row=row, column=1, value=row-2)
    worksheet.cell(row=row, column=2, value=total_cost)
    worksheet.cell(row=row, column=3, value=total_time)
    worksheet.cell(row=row, column=4, value=htn_total_time)
    worksheet.cell(row=row, column=5, value=step_size)

    row += 1

# Save the Excel workbook
workbook.save('output.xlsx')